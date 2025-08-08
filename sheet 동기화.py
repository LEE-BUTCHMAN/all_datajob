import gspread
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook

# ⭐️ 여기만 바꾸면 됨!
TARGET_WEEK = 31  # 원하는 주차 입력

# 정확한 OneDrive 경로
EXCEL_PATH = '/Users/sfn/Library/CloudStorage/OneDrive-에쓰푸드/에쓰푸드의 파일 - FreshOn Market팀/butchman/작업중/automation_excel.xlsx'

# Google Sheets 인증
scope = ['https://spreadsheets.google.com/feeds',
          'https://www.googleapis.com/auth/drive']
creds = Credentials.from_service_account_file(
    '/Users/sfn/Downloads/automation-data-467003-6310e37f0e5c.json',
    scopes=scope
)
gc = gspread.authorize(creds)

# Google Sheets 데이터 읽기
sheet = gc.open_by_key('1zmujGEM6C51LxrljTlIsKAwxgXAj82K9YfkQxpg7OjE')
worksheet = sheet.worksheet('automation(매출)')
all_data = worksheet.get_all_values()

# 지정한 주차 찾기
week_row = all_data[1]
target_col = None

for idx, cell in enumerate(week_row):
    if str(TARGET_WEEK) in str(cell):
        target_col = idx
        print(f"📅 {TARGET_WEEK}주차 동기화 시작")
        break

if target_col is None:
    print(f"❌ {TARGET_WEEK}주차를 찾을 수 없습니다.")
    exit()

# 해당 주차 데이터 추출
sync_data = []
for row in all_data:
    if len(row) > target_col:
        sync_data.append([row[0], row[target_col]])

# Excel 업데이트
wb = load_workbook(EXCEL_PATH)

if 'test' not in wb.sheetnames:
    wb.create_sheet('test')

ws = wb['test']

for row in ws.iter_rows():
    for cell in row:
        cell.value = None

for row_idx, row_data in enumerate(sync_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

wb.save(EXCEL_PATH)
wb.close()

print(f"✅ {TARGET_WEEK}주차 동기화 완료!")
print("📁 OneDrive에 자동 업로드됨")